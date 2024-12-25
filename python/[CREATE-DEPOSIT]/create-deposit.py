import argparse
import requests
import time
import random
from colorama import Fore
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import os
import sys
import io
import threading

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

blue = Fore.BLUE
green = Fore.GREEN
yellow = Fore.YELLOW
red = Fore.RED
roxo = Fore.MAGENTA
reset = Fore.RESET

config_path = './config.json'
if not os.path.exists(config_path):
    print(f"{red}[ERROR]{reset} {roxo}O arquivo de configuração não foi encontrado.{reset}", end='')
    sys.stdout.flush()
    sys.exit(1)

with open(config_path) as config_file:
    config = json.load(config_file)

profiles_create_path = './[UNKOYNX7]/[CREATE-DEPOSIT]/profilesCreate.txt'
accountCreatePath = './[UNKOYNX7]/[CREATE-DEPOSIT]/accountCreateDeposit.txt'
proxiesUsedPath = './[UNKOYNX7]/[CREATE-DEPOSIT]/proxiesUsed.txt'
proxys_path = './[UNKOYNX7]/[CREATE-DEPOSIT]/proxys.txt'


def get_accounts():
    if not os.path.exists(profiles_create_path):
        print(f"{red}[ERROR]{reset} {roxo}O arquivo de perfils criados não foi encontrado.{reset}", end='')
        sys.stdout.flush()
        sys.exit(1)

    with open(profiles_create_path, 'r', encoding='utf-8') as file:
        accounts = file.readlines()
    
    accounts = [line.strip() for line in accounts if line.strip()]
    
    if not accounts:
        print(f"{red}[ERROR]{reset} {roxo}Nenhuma conta disponível.{reset}")
        sys.stdout.flush()
        return []

    return accounts

def remove_account_line(account_line):
    with open(profiles_create_path, 'r', encoding='utf-8') as file:
        accounts = file.readlines()
    updated_accounts = [account.strip() for account in accounts if account.strip() and account.strip() != account_line]
    with open(profiles_create_path, 'w', encoding='utf-8') as file:
        file.write('\n'.join(updated_accounts) + '\n' if updated_accounts else '')

    sys.stdout.flush()

def click_button_when_clickable(driver, cpf, xpath):
    try:
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        button.click()
        print(f"{yellow}[INFO]{reset} {roxo}({cpf}) - Botão clicado com sucesso!{reset}", end='')
        sys.stdout.flush()
    except Exception as e:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Erro ao tentar clicar no botão.{reset}", end='')
        sys.stdout.flush()

def type_with_delay(element, text, delay=100):
    element.click()
    for char in text:
        element.send_keys(char)
        time.sleep(delay / 1000.0)

def save_proxies(proxyIP, proxyPort, proxyUsername, proxyPassword):
    proxy_line = f"{proxyIP}:{proxyPort}:{proxyUsername}:{proxyPassword}"

    existing_proxies = set()
    if os.path.exists(proxys_path):
        with open(proxys_path, 'r', encoding='utf-8') as file:
            existing_proxies = {line.strip() for line in file if line.strip()}

    if proxy_line and proxy_line not in existing_proxies:
        existing_proxies.add(proxy_line)

        with open(proxys_path, 'w', encoding='utf-8') as file:
            file_content = '\n'.join(sorted(existing_proxies))
            file.write(file_content)

def fill_form_and_submit(driver, account_data, email, password, cpf, firstName, lastName, proxyIP, proxyPort, proxyUsername, proxyPassword, cep_text, endereco_text, cidade_text, estado_text, profile_id):
    try:
        email_container = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="auth-modal"]/div[2]/form/div[1]/div'))
        )
        email_field = email_container.find_element(By.XPATH, './input')
        email_field.send_keys(email)
        
        senha_container = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="auth-modal"]/div[2]/form/div[2]/div'))
        )
        senha_field = senha_container.find_element(By.XPATH, './input')
        senha_field.send_keys(password)

        cpf_field = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="auth-modal"]/div[2]/form/div[3]/div/div/div/div/input'))
        )
        type_with_delay(cpf_field, cpf, delay=100)

        try:
            error_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "div.error-label"))
            )
            error_message = error_element.text
            if ("Este CPF pertence a um menor e não pode ser aceito" in error_message or
                "O CPF fornecido é inválido. Insira um número de CPF válido." in error_message):
                print(f"{red}[ERROR]{reset} {roxo}({cpf}) - {error_message}{reset}")
                save_proxies(proxyIP, proxyPort, proxyUsername, proxyPassword)
                delete_profile(profile_id, account_data, cpf)
                return
        except Exception:
            pass

        click_button_when_clickable(driver, cpf, '//*[@id="auth-modal"]/div[2]/form/div[5]/button')

        try:
            cpf_error_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="auth-modal"]/div[2]/form/div[3]/div/div/div/div[2]'))
            )
            if cpf_error_element.is_displayed():
                print(f"{red}[ERROR]{reset} {roxo}({cpf}) - O cpf já está cadastrado.{reset}")
                save_proxies(proxyIP, proxyPort, proxyUsername, proxyPassword)
                delete_profile(profile_id, account_data, cpf)
                return
        except Exception:
            pass

        try:
            email_error_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="auth-modal"]/div[2]/form/div[2]'))
            )
            if email_error_element:
                print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Já existe uma conta com esse e-mail.{reset}")
                save_proxies(proxyIP, proxyPort, proxyUsername, proxyPassword)
                delete_profile(profile_id, account_data, cpf)
                return
        except Exception:
            pass

            success_message = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//div[@class="welcome-header" and text()="Bem-vindo ao Jonbet!"]'))
            )
            if success_message:
                print(f"{green}[SUCCESS]{reset} {roxo}({cpf}) - Conta criada com sucesso!{reset}", end='')
                sys.stdout.flush()

                click_button_when_clickable(driver, cpf, '//*[@id="parent-modal-close"]')
                time.sleep(1)

                click_button_when_clickable(driver, cpf, '//*[@id="header-deposit"]')

                click_button_when_clickable(driver, cpf, '//*[@id="bonus-selector"]/button')
                time.sleep(0.5)

                click_button_when_clickable(driver, cpf, '//*[@id="method-selector"]/div/div[1]')

                firstName_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="new-transaction"]/div/div[2]/div[1]/div/div[2]/div/div/div'))
                )
                firstName_field = firstName_container.find_element(By.XPATH, './input')
                firstName_field.send_keys(firstName)

                lastName_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="new-transaction"]/div/div[2]/div[1]/div/div[3]/div/div/div'))
                )
                lastName_field = lastName_container.find_element(By.XPATH, './input')
                lastName_field.send_keys(lastName)

                random_value = round(random.uniform(41, 45), 1)

                random_value_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="new-transaction"]/div/div[2]/div[1]/div/div[4]/div/div/div[2]/div[1]'))
                )
                random_value_field = random_value_container.find_element(By.XPATH, './input')
                random_value_field.send_keys(random_value)

                click_button_when_clickable(driver, cpf, '//*[@id="new-transaction"]/div/div[2]/div[2]/div/button[2]')

                copy_element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//*[@id="new-transaction"]/div/div[2]/div[1]/div[2]/div[1]'))
                )
                copy_text = copy_element.text
                print(f"{green}[SUCCESS]{reset} {roxo}[{random_value}] ({cpf}) - Pagamento gerado com sucesso!{reset}", end='')
                sys.stdout.flush()

                driver.get("https://jonbet.com/pt/account/profile/personal")
                time.sleep(1.5)

                endereco_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="user-info"]/div/div[5]/div/div/div'))
                )
                endereco_field = endereco_container.find_element(By.XPATH, './input')
                endereco_field.send_keys(endereco_text)

                cep_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="user-info"]/div/div[9]/div/div/div'))
                )
                cep_field = cep_container.find_element(By.XPATH, './input')
                cep_field.send_keys(cep_text)

                cidade_container = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="user-info"]/div/div[8]/div/div/div'))
                )
                cidade_field = cidade_container.find_element(By.XPATH, './input')
                cidade_field.send_keys(cidade_text)

                click_button_when_clickable(driver, cpf, '//*[@id="user-info"]/div/div[6]/div/div/a/div/div')

                click_button_when_clickable(driver, cpf, '//*[@id="user-info"]/div/div[6]/div/div/div/div/div[25]')

                estados_xpaths = {
                    "AC": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[1]",
                    "AL": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[2]",
                    "AP": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[3]",
                    "AM": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[4]",
                    "BA": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[5]",
                    "CE": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[6]",
                    "DF": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[7]",
                    "ES": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[8]",
                    "GO": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[9]",
                    "MA": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[10]",
                    "MT": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[11]",
                    "MS": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[12]",
                    "MG": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[13]",
                    "PA": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[14]",
                    "PB": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[15]",
                    "PR": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[16]",
                    "PE": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[17]",
                    "PI": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[18]",
                    "RJ": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[19]",
                    "RN": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[20]",
                    "RS": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[21]",
                    "RO": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[22]",
                    "RR": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[23]",
                    "SC": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[24]",
                    "SP": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[25]",
                    "SE": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[26]",
                    "TO": "//*[@id='user-info']/div/div[7]/div/div/div/div/div[27]"
                }

                def get_xpath_estado(estado_text):
                    return estados_xpaths.get(estado_text, None)
                
                click_button_when_clickable(driver, cpf, '//*[@id="user-info"]/div/div[7]/div/div/a/div/div')

                xpath_estado = get_xpath_estado(estado_text)
                click_button_when_clickable(driver, cpf, f'{xpath_estado}')

                click_button_when_clickable(driver, cpf, '//*[@id="user-info"]/button')
                save_account(firstName, lastName, email, password, cpf, proxyIP, proxyPort, proxyUsername, proxyPassword, copy_text, random_value)
                delete_profile(profile_id, account_data, cpf)

    except Exception as e:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Erro ao criar a conta e deposito.{reset}", end='')
        sys.stdout.flush()

def check_if_content_is_visible(driver):
    try:
        content_element = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="policy-regulation-popup"]/div/div[2]/div/button'))
        )
        return True
    except Exception:
        return False
    
def save_account(firstName, lastName, email, password, cpf, proxyIP, proxyPort, proxyUsername, proxyPassword, copy_text, random_value):
    excel_path = './[UNKOYNX7]/[CREATE-DEPOSIT]/accountsDeposit.xlsx'
    
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    if not os.path.exists(excel_path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Accounts"

        headers = ["Primeiro Nome", "Ultimo Nome", "Email", "Password", "CPF", "Proxy", "Valor", "Copia & Cola"]
        worksheet.append(headers)

        workbook.save(excel_path)
    else:
        workbook = load_workbook(excel_path)
        worksheet = workbook.active

    proxy = f"{proxyIP}:{proxyPort}:{proxyUsername}:{proxyPassword}"

    account_data = [firstName, lastName, email, password, cpf, proxy, random_value, copy_text]
    worksheet.append(account_data)

    workbook.save(excel_path)

def create_account(account_data):
    if not account_data or not isinstance(account_data, str):
        print(f"{red}[ERROR]{reset} {roxo}Dados da conta não são válidos.{reset}", end='')
        sys.stdout.flush()
        return

    credentials, proxy, cep_info, additional_info = account_data.split(' - ')
    email, password, cpf, firstName, lastName = credentials.split(':')
    proxyIP, proxyPort, proxyUsername, proxyPassword = proxy.split(':')
    cep_text, endereco_text, cidade_text, estado_text = cep_info.split(':')
    profile_id = additional_info.split(':')[0].strip()

    open_url = f"{config['AdsPower']['urlLocal']}/api/v1/browser/start?user_id={profile_id}"

    resp = requests.get(open_url).json()

    if resp["code"] != 0:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Perfil no adsPower não existe.{reset}", end='')
        sys.stdout.flush()
        return

    chrome_driver = resp["data"]["webdriver"]
    debugger_address = resp["data"]["ws"]["selenium"]

    if ':' not in debugger_address:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Ocorreu um erro no debuggerAddress, contate aos administradores!'.{reset}", end='')
        sys.stdout.flush()
        return
    else:
        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", debugger_address)

    driver = None

    try:
        driver = webdriver.Chrome(service=Service(chrome_driver), options=chrome_options)
        print(f"{green}[SUCCESS]{reset} {roxo}({cpf}) - Conectado ao navegador com sucesso!{reset}", end='')
        sys.stdout.flush()

        time.sleep(2)

        driver.get(f"{config['AdsPower']['linkJonbet']}")
        print(f"{yellow}[INFO]{reset} {roxo}({cpf}) - Link aberto com sucesso!{reset}", end='')
        sys.stdout.flush()

        if check_if_content_is_visible(driver):
            click_button_when_clickable(driver, cpf, '//*[@id="policy-regulation-popup"]/div/div[2]/div/button')
        else:
            print(f"{yellow}[INFO]{reset} {roxo}({cpf}) - Botão de cookies já aceito.{reset}", end='')
            sys.stdout.flush()

        click_button_when_clickable(driver, cpf, '//*[@id="landing-home"]/div[1]/div/div/div[1]/div[2]/div/button')

        fill_form_and_submit(driver, account_data, email, password, cpf, firstName, lastName, proxyIP, proxyPort, proxyUsername, proxyPassword, cep_text, endereco_text, cidade_text, estado_text, profile_id)

    except Exception as error:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Ocorreu um erro conectar ao selenium.{reset}\n{error}", end='')
        sys.stdout.flush()
    finally:
        if driver:
            driver.quit()

def close_browser(profile_id, cpf):
    close_url = f"{config['AdsPower']['urlLocal']}/api/v1/browser/stop?user_id={profile_id}"
    requests.get(close_url)
    print(f"{green}[SUCCESS]{reset} {roxo}({cpf}) - Navegador fechado com sucesso!{reset}", end='')
    sys.stdout.flush()

def check_browser_status(profile_id, cpf):
    url = f"{config['AdsPower']['urlLocal']}/api/v1/browser/active"
    params = {"user_id": profile_id}

    while True:
        try:
            response = requests.get(url, params=params)
            data = response.json()
            if data["code"] == 0:
                if data["data"]["status"] == "Inactive":
                    break
            else:
                print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Falha ao verificar o status do navegador.{reset}", end='')
                sys.stdout.flush()
                break
        except Exception as error:
            print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Erro ao verificar o status do navegador.{reset}", end='')
            sys.stdout.flush()
            break

        time.sleep(1)

def delete_profile(profile_id, account_data, cpf):
    close_browser(profile_id, cpf)
    check_browser_status(profile_id, cpf)
    time.sleep(1)

    url = f"{config['AdsPower']['urlLocal']}/api/v1/user/delete"
    params = {"user_ids": [profile_id]}

    try:
        response = requests.post(url, json=params)
        data = response.json()
        if data["code"] == 0:
            remove_account_line(account_data)
        elif "is being used by" in data.get("msg", ""):
            pass
        else:
            pass
    except Exception as error:
        print(f"{red}[ERROR]{reset} {roxo}({cpf}) - Erro ao tentar deletar perfil.{reset}")
        sys.stdout.flush()

def parse_arguments():
    parser = argparse.ArgumentParser(description='Script para criar contas com múltiplos navegadores.')
    parser.add_argument('--max_browsers', type=int, required=True, help='Número máximo de navegadores a serem usados.')
    return parser.parse_args()

def main():
    args = parse_arguments()
    accounts = get_accounts()
    
    threads = []
    max_browsers = min(args.max_browsers, len(accounts))

    for i, account in enumerate(accounts[:max_browsers]):
        if i > 0:
            time.sleep(0.500)

        thread = threading.Thread(target=create_account, args=(account,))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

if __name__ == "__main__":
    main()